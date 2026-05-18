"""
실습 2: CSV/Excel → Chroma → Q&A 봇
핵심 개념: Load → Split → Embed → Store → Retrieve → Generate
실행: python 02_csv_excel_rag.py
준비: data/ 디렉토리에 CSV 또는 Excel 파일 배치
"""

import csv
import glob
import os

from langchain.schema import Document


def load_csv(path: str, encoding: str = "utf-8-sig") -> list[Document]:
    docs = []
    with open(path, encoding=encoding) as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            content = "\n".join(f"{k}: {v}" for k, v in row.items() if v)
            if content.strip():
                docs.append(Document(
                    page_content=content,
                    metadata={"source": os.path.basename(path), "row": i},
                ))
    return docs


def load_excel(path: str) -> list[Document]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    docs = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
        for i, row in enumerate(rows[1:], start=1):
            content = "\n".join(
                f"{headers[j]}: {v}" for j, v in enumerate(row) if v is not None
            )
            if content.strip():
                docs.append(Document(
                    page_content=content,
                    metadata={"source": os.path.basename(path), "sheet": sheet_name, "row": i},
                ))
    return docs


def load_data_dir(data_dir: str) -> list[Document]:
    all_docs = []
    patterns = ["*.csv", "*.xlsx", "*.xls"]
    for pattern in patterns:
        for path in glob.glob(os.path.join(data_dir, pattern)):
            if path.endswith(".csv"):
                all_docs.extend(load_csv(path))
            else:
                all_docs.extend(load_excel(path))
    return all_docs


def main():
    from dotenv import load_dotenv
    from langchain.chains import RetrievalQA
    from langchain.prompts import PromptTemplate
    from langchain.text_splitter import RecursiveCharacterTextSplitter
    from langchain_chroma import Chroma
    from langchain_huggingface import HuggingFaceEmbeddings
    from langchain_openai import ChatOpenAI

    load_dotenv()

    DATA_DIR = "./data"
    PERSIST_DIR = "./chroma_db"

    all_docs = load_data_dir(DATA_DIR)
    if not all_docs:
        raise FileNotFoundError(f"{DATA_DIR}/ 에 CSV 또는 Excel 파일이 없습니다.")

    chunks = RecursiveCharacterTextSplitter(
        chunk_size=500,
        chunk_overlap=50,
        separators=["\n\n", "\n", " "],
    ).split_documents(all_docs)
    print(f"[Index] 문서 {len(all_docs)}행 → 청크 {len(chunks)}개")

    # EMBED_CACHE_DIR: 수동 설치한 모델 경로 (python 실행 디렉토리 기준)
    embeddings = HuggingFaceEmbeddings(
        model_name=os.getenv("EMBED_MODEL", "BAAI/bge-m3"),
        cache_folder=os.getenv("EMBED_CACHE_DIR", "./models"),
    )

    if os.path.exists(PERSIST_DIR) and os.listdir(PERSIST_DIR):
        vectorstore = Chroma(persist_directory=PERSIST_DIR, embedding_function=embeddings)
        print("[Index] 기존 Chroma DB 재사용 (임베딩 재계산 안 함)")
    else:
        vectorstore = Chroma.from_documents(chunks, embeddings, persist_directory=PERSIST_DIR)
        print("[Index] 새 Chroma DB 생성 + 임베딩 저장")

    prompt = PromptTemplate(
        template="""다음 컨텍스트만 사용하여 한국어로 답하세요.

컨텍스트:
{context}

질문: {question}

규칙:
- 컨텍스트에 없으면 "데이터에서 찾을 수 없습니다."라고만 답하세요.
- 답변 끝에 [출처: 파일명, 행 번호] 형식으로 근거를 표시하세요.

답변:""",
        input_variables=["context", "question"],
    )

    qa = RetrievalQA.from_chain_type(
        llm=ChatOpenAI(
            base_url=os.getenv("LLM_BASE_URL"),
            api_key=os.getenv("LLM_API_KEY"),
            model=os.getenv("LLM_MODEL"),
            temperature=0,
        ),
        chain_type="stuff",
        retriever=vectorstore.as_retriever(search_kwargs={"k": 3}),
        chain_type_kwargs={"prompt": prompt},
        return_source_documents=True,
    )

    print("\nCSV/Excel Q&A 봇 준비 완료. 질문을 입력하세요.")
    while True:
        q = input("\n질문 (exit 입력 시 종료): ").strip()
        if q.lower() in {"exit", "quit", ""}:
            break
        out = qa.invoke({"query": q})
        print(f"\n[답변]\n{out['result']}")
        print("\n[근거]")
        for doc in out["source_documents"]:
            meta = doc.metadata
            loc = f"row {meta.get('row', '?')}"
            if "sheet" in meta:
                loc = f"sheet={meta['sheet']} {loc}"
            print(f"  - {meta.get('source', '?')} ({loc}): {doc.page_content[:80].replace(chr(10), ' ')}...")


if __name__ == "__main__":
    main()
