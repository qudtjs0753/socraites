"""
파일 로더 테스트 — CSV / Excel / TXT / MD / LOG / 이미지(OCR)
실행: pytest test_loaders.py -v
"""

import csv
import os

import pytest

from rag_pipeline import load_csv, load_excel, load_text, load_image, load_data_dir


# ── 헬퍼 ─────────────────────────────────────────────────────────────────────

def write_csv(path, rows: list[dict], encoding="utf-8-sig"):
    with open(path, "w", encoding=encoding, newline="") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)


def write_excel(path, sheets: dict[str, list[dict]]):
    import openpyxl
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet(sheet_name)
        if first:
            ws.title = sheet_name
            first = False
        ws.append(list(rows[0].keys()))
        for row in rows:
            ws.append(list(row.values()))
    wb.save(path)


# ── CSV ───────────────────────────────────────────────────────────────────────

class TestLoadCsv:
    def test_기본_행이_Document로_변환된다(self, tmp_path):
        path = tmp_path / "sample.csv"
        write_csv(path, [
            {"이름": "홍길동", "부서": "개발팀", "직급": "과장"},
            {"이름": "김철수", "부서": "인프라팀", "직급": "대리"},
        ])
        docs = load_csv(str(path))
        assert len(docs) == 2
        assert "홍길동" in docs[0].page_content
        assert "개발팀" in docs[0].page_content

    def test_메타데이터에_파일명과_행번호가_들어간다(self, tmp_path):
        path = tmp_path / "meta.csv"
        write_csv(path, [{"항목": "A", "값": "1"}, {"항목": "B", "값": "2"}])
        docs = load_csv(str(path))
        assert docs[0].metadata["source"] == "meta.csv"
        assert docs[0].metadata["row"] == 0
        assert docs[1].metadata["row"] == 1

    def test_빈_값_컬럼은_content에서_제외된다(self, tmp_path):
        path = tmp_path / "empty_col.csv"
        write_csv(path, [{"이름": "홍길동", "메모": "", "부서": "개발팀"}])
        docs = load_csv(str(path))
        assert "메모" not in docs[0].page_content
        assert "이름: 홍길동" in docs[0].page_content

    def test_모든_컬럼이_빈_행은_제외된다(self, tmp_path):
        path = tmp_path / "all_empty.csv"
        write_csv(path, [{"이름": "홍길동", "부서": "개발팀"}, {"이름": "", "부서": ""}])
        docs = load_csv(str(path))
        assert len(docs) == 1

    def test_content_형식이_key_value_줄바꿈이다(self, tmp_path):
        path = tmp_path / "format.csv"
        write_csv(path, [{"이름": "홍길동", "부서": "개발팀"}])
        docs = load_csv(str(path))
        assert docs[0].page_content == "이름: 홍길동\n부서: 개발팀"

    def test_utf8_bom_없는_파일도_읽힌다(self, tmp_path):
        path = tmp_path / "utf8.csv"
        write_csv(path, [{"이름": "홍길동"}], encoding="utf-8")
        docs = load_csv(str(path), encoding="utf-8")
        assert "홍길동" in docs[0].page_content

    def test_euc_kr_파일도_encoding_지정하면_읽힌다(self, tmp_path):
        path = tmp_path / "euckr.csv"
        write_csv(path, [{"이름": "홍길동", "부서": "개발팀"}], encoding="euc-kr")
        docs = load_csv(str(path), encoding="euc-kr")
        assert "홍길동" in docs[0].page_content

    def test_헤더만_있는_빈_파일은_빈_리스트를_반환한다(self, tmp_path):
        path = tmp_path / "empty.csv"
        path.write_text("이름,부서\n", encoding="utf-8-sig")
        assert load_csv(str(path)) == []

    def test_파일이_없으면_예외가_발생한다(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            load_csv(str(tmp_path / "없는파일.csv"))


# ── Excel ─────────────────────────────────────────────────────────────────────

class TestLoadExcel:
    def test_단일_시트_행이_Document로_변환된다(self, tmp_path):
        path = tmp_path / "sample.xlsx"
        write_excel(path, {"Sheet1": [
            {"이름": "홍길동", "부서": "개발팀"},
            {"이름": "김철수", "부서": "인프라팀"},
        ]})
        docs = load_excel(str(path))
        assert len(docs) == 2
        assert "홍길동" in docs[0].page_content

    def test_메타데이터에_파일명_시트명_행번호가_들어간다(self, tmp_path):
        path = tmp_path / "meta.xlsx"
        write_excel(path, {"직원": [{"이름": "홍길동", "직급": "과장"}]})
        docs = load_excel(str(path))
        assert docs[0].metadata["source"] == "meta.xlsx"
        assert docs[0].metadata["sheet"] == "직원"
        assert docs[0].metadata["row"] == 1

    def test_다중_시트를_모두_읽는다(self, tmp_path):
        path = tmp_path / "multi.xlsx"
        write_excel(path, {
            "개발팀": [{"이름": "홍길동"}, {"이름": "이영희"}],
            "인프라팀": [{"이름": "김철수"}],
        })
        docs = load_excel(str(path))
        assert len(docs) == 3
        sheets = {d.metadata["sheet"] for d in docs}
        assert sheets == {"개발팀", "인프라팀"}

    def test_None_셀은_content에서_제외된다(self, tmp_path):
        path = tmp_path / "none_cell.xlsx"
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["이름", "메모", "부서"])
        ws.append(["홍길동", None, "개발팀"])
        wb.save(path)
        docs = load_excel(str(path))
        assert "메모" not in docs[0].page_content
        assert "이름: 홍길동" in docs[0].page_content

    def test_빈_시트는_건너뛴다(self, tmp_path):
        path = tmp_path / "empty_sheet.xlsx"
        import openpyxl
        wb = openpyxl.Workbook()
        wb.active.title = "빈시트"
        wb.create_sheet("데이터").append(["이름"])
        wb["데이터"].append(["홍길동"])
        wb.save(path)
        docs = load_excel(str(path))
        assert all(d.metadata["sheet"] == "데이터" for d in docs)


# ── TXT / MD / LOG ────────────────────────────────────────────────────────────

class TestLoadText:
    @pytest.mark.parametrize("ext", [".txt", ".md", ".log"])
    def test_파일_전체가_하나의_Document로_로드된다(self, tmp_path, ext):
        path = tmp_path / f"sample{ext}"
        path.write_text("첫 번째 줄\n두 번째 줄\n세 번째 줄", encoding="utf-8")
        docs = load_text(str(path))
        assert len(docs) == 1
        assert "첫 번째 줄" in docs[0].page_content
        assert "세 번째 줄" in docs[0].page_content

    def test_메타데이터에_파일명이_들어간다(self, tmp_path):
        path = tmp_path / "report.md"
        path.write_text("# 보고서", encoding="utf-8")
        docs = load_text(str(path))
        assert docs[0].metadata["source"] == "report.md"

    def test_cp949_파일도_읽힌다(self, tmp_path):
        path = tmp_path / "legacy.txt"
        path.write_bytes("한글 내용".encode("cp949"))
        docs = load_text(str(path))
        assert "한글 내용" in docs[0].page_content

    def test_빈_파일은_빈_리스트를_반환한다(self, tmp_path):
        path = tmp_path / "empty.txt"
        path.write_text("", encoding="utf-8")
        assert load_text(str(path)) == []

    def test_공백만_있는_파일은_빈_리스트를_반환한다(self, tmp_path):
        path = tmp_path / "whitespace.log"
        path.write_text("   \n\n  ", encoding="utf-8")
        assert load_text(str(path)) == []


# ── 이미지(OCR) ───────────────────────────────────────────────────────────────

pytesseract_available = pytest.mark.skipif(
    not __import__("importlib").util.find_spec("pytesseract"),
    reason="pytesseract 미설치"
)


class TestLoadImage:
    @pytesseract_available
    def test_텍스트가_있는_이미지에서_Document를_반환한다(self, tmp_path):
        from PIL import Image, ImageDraw
        img = Image.new("RGB", (200, 50), color="white")
        ImageDraw.Draw(img).text((10, 10), "hello world", fill="black")
        path = tmp_path / "text.png"
        img.save(path)
        docs = load_image(str(path))
        assert len(docs) == 1
        assert docs[0].metadata["source"] == "text.png"

    @pytesseract_available
    def test_텍스트_없는_이미지는_빈_리스트를_반환한다(self, tmp_path):
        from PIL import Image
        img = Image.new("RGB", (100, 100), color="white")
        path = tmp_path / "blank.jpg"
        img.save(path)
        docs = load_image(str(path))
        assert docs == []

    def test_pytesseract_미설치시_ImportError가_발생한다(self, tmp_path, monkeypatch):
        import builtins
        real_import = builtins.__import__

        def mock_import(name, *args, **kwargs):
            if name in ("pytesseract", "PIL"):
                raise ImportError
            return real_import(name, *args, **kwargs)

        from PIL import Image
        img = Image.new("RGB", (10, 10))
        path = tmp_path / "img.png"
        img.save(path)

        monkeypatch.setattr(builtins, "__import__", mock_import)
        with pytest.raises(ImportError):
            load_image(str(path))


# ── load_data_dir 통합 ────────────────────────────────────────────────────────

class TestLoadDataDir:
    def test_혼합_파일을_모두_읽는다(self, tmp_path):
        write_csv(tmp_path / "data.csv", [{"항목": "CSV데이터"}])
        (tmp_path / "note.txt").write_text("텍스트 내용", encoding="utf-8")
        (tmp_path / "log.log").write_text("로그 내용", encoding="utf-8")

        docs = load_data_dir(str(tmp_path))

        sources = {d.metadata["source"] for d in docs}
        assert "data.csv" in sources
        assert "note.txt" in sources
        assert "log.log" in sources

    def test_지원하지_않는_확장자는_무시한다(self, tmp_path):
        (tmp_path / "ignore.pdf").write_bytes(b"%PDF-1.4")
        (tmp_path / "note.txt").write_text("내용", encoding="utf-8")

        docs = load_data_dir(str(tmp_path))

        sources = {d.metadata["source"] for d in docs}
        assert "ignore.pdf" not in sources
        assert "note.txt" in sources

    def test_빈_디렉토리는_빈_리스트를_반환한다(self, tmp_path):
        assert load_data_dir(str(tmp_path)) == []
