"""
실습 2: 다중 파일 형식 → Chroma → Q&A 봇
지원 형식: CSV, Excel(.xlsx/.xls), TXT, MD, LOG, JPG/PNG(OCR)
실행: python rag_pipeline.py
준비: data/ 디렉토리에 지원 파일 배치
"""

import csv
import glob
import os
from pathlib import Path

from langchain.schema import Document


# ── 파일별 로더 ─────────────────────────────────────────────────────────────

def load_csv(path: str, encoding: str = "utf-8-sig") -> list[Document]:
    docs = []
    with open(path, encoding=encoding) as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            content = "\n".join(f"{k}: {v}" for k, v in row.items() if v)
            if content.strip():
                docs.append(Document(
                    page_content=content,
                    metadata={"source": os.path.basename(path), "row": i},
                ))
    return docs


def load_excel(path: str) -> list[Document]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    docs = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
        for i, row in enumerate(rows[1:], start=1):
            content = "\n".join(
                f"{headers[j]}: {v}" for j, v in enumerate(row) if v is not None
            )
            if content.strip():
                docs.append(Document(
                    page_content=content,
                    metadata={"source": os.path.basename(path), "sheet": sheet_name, "row": i},
                ))
    return docs


def load_text(path: str, encoding: str = "utf-8") -> list[Document]:
    """TXT / MD / LOG 공용 로더 — 파일 전체를 하나의 Document로 로드"""
    try:
        text = Path(path).read_text(encoding=encoding)
    except UnicodeDecodeError:
        text = Path(path).read_text(encoding="cp949", errors="replace")
    if not text.strip():
        return []
    return [Document(
        page_content=text,
        metadata={"source": os.path.basename(path)},
    )]


def load_image(path: str) -> list[Document]:
    """JPG / PNG → OCR(pytesseract)로 텍스트 추출"""
    try:
        from PIL import Image
        import pytesseract
    except ImportError:
        raise ImportError("pip install pillow pytesseract  # 그리고 OS에 tesseract 설치 필요")
    text = pytesseract.image_to_string(Image.open(path), lang="kor+eng")
    if not text.strip():
        return []
    return [Document(
        page_content=text,
        metadata={"source": os.path.basename(path)},
    )]


# ── 디렉토리 일괄 로드 ────────────────────────────────────────────────────────

LOADERS = {
    ".csv":  load_csv,
    ".xlsx": load_excel,
    ".xls":  load_excel,
    ".txt":  load_text,
    ".md":   load_text,
    ".log":  load_text,
    ".jpg":  load_image,
    ".jpeg": load_image,
    ".png":  load_image,
}


def load_data_dir(data_dir: str) -> list[Document]:
    all_docs = []
    for ext, loader in LOADERS.items():
        for path in glob.glob(os.path.join(data_dir, f"*{ext}")):
            docs = loader(path)
            all_docs.extend(docs)
            print(f"  [Load] {os.path.basename(path)} → {len(docs)}개 청크 예정")
    return all_docs


# ── 메인 ──────────────────────────────────────────────────────────────────────

def main():
    import requests
    from typing import List, Any
    from dotenv import load_dotenv
    from langchain.chains import RetrievalQA
    from langchain.prompts import PromptTemplate
    from langchain.text_splitter import RecursiveCharacterTextSplitter
    from langchain_chroma import Chroma
    from langchain_huggingface import HuggingFaceEmbeddings
    from langchain_core.language_models.chat_models import BaseChatModel
    from langchain_core.messages import AIMessage, HumanMessage, SystemMessage, BaseMessage
    from langchain_core.outputs import ChatGeneration, ChatResult

    load_dotenv()

    class InternalChatLLM(BaseChatModel):
        base_url: str = ""
        api_key: str = ""
        model_name: str = ""
        temperature: float = 0

        def _generate(self, messages: List[BaseMessage], stop=None, run_manager=None, **kwargs) -> ChatResult:
            role_map = {HumanMessage: "user", SystemMessage: "system", AIMessage: "assistant"}
            resp = requests.post(
                f"{self.base_url}/chat/completions",
                json={
                    "model": self.model_name,
                    "messages": [{"role": role_map.get(type(m), "user"), "content": m.content} for m in messages],
                    "temperature": self.temperature,
                },
                headers={"Authorization": f"Bearer {self.api_key}"},
                timeout=60,
            )
            resp.raise_for_status()
            return ChatResult(generations=[ChatGeneration(
                message=AIMessage(content=resp.json()["choices"][0]["message"]["content"])
            )])

        @property
        def _llm_type(self) -> str:
            return "internal_llm"

    DATA_DIR = "./data"
    PERSIST_DIR = "./chroma_db"

    print("[Load] data/ 디렉토리 스캔 중...")
    all_docs = load_data_dir(DATA_DIR)
    if not all_docs:
        raise FileNotFoundError(
            f"{DATA_DIR}/ 에 지원 파일(CSV/Excel/TXT/MD/LOG)이 없습니다."
        )

    chunks = RecursiveCharacterTextSplitter(
        chunk_size=500,
        chunk_overlap=50,
        separators=["\n\n", "\n", " "],
    ).split_documents(all_docs)
    print(f"[Split] 문서 {len(all_docs)}개 → 청크 {len(chunks)}개")

    embeddings = HuggingFaceEmbeddings(
        model_name=os.getenv("EMBED_MODEL", "BAAI/bge-m3"),
        cache_folder=os.getenv("EMBED_CACHE_DIR", "./hf_cache"),
        model_kwargs={"local_files_only": True},
        encode_kwargs={"normalize_embeddings": True},
    )

    if os.path.exists(PERSIST_DIR) and os.listdir(PERSIST_DIR):
        vectorstore = Chroma(persist_directory=PERSIST_DIR, embedding_function=embeddings)
        print("[Index] 기존 Chroma DB 재사용")
    else:
        vectorstore = Chroma.from_documents(chunks, embeddings, persist_directory=PERSIST_DIR)
        print("[Index] 새 Chroma DB 생성 + 임베딩 저장")

    llm = InternalChatLLM(
        base_url=os.getenv("LLM_BASE_URL", ""),
        api_key=os.getenv("LLM_API_KEY", ""),
        model_name=os.getenv("LLM_MODEL", ""),
        temperature=0,
    )

    prompt = PromptTemplate(
        template="""다음 컨텍스트만 사용하여 한국어로 답하세요.

컨텍스트:
{context}

질문: {question}

규칙:
- 컨텍스트에 없으면 "데이터에서 찾을 수 없습니다."라고만 답하세요.
- 답변 끝에 [출처: 파일명] 형식으로 근거를 표시하세요.

답변:""",
        input_variables=["context", "question"],
    )

    qa = RetrievalQA.from_chain_type(
        llm=llm,
        chain_type="stuff",
        retriever=vectorstore.as_retriever(search_kwargs={"k": 3}),
        chain_type_kwargs={"prompt": prompt},
        return_source_documents=True,
    )

    print("\nQ&A 봇 준비 완료. 질문을 입력하세요.")
    while True:
        q = input("\n질문 (exit 입력 시 종료): ").strip()
        if q.lower() in {"exit", "quit", ""}:
            break
        out = qa.invoke({"query": q})
        print(f"\n[답변]\n{out['result']}")
        print("\n[근거]")
        for doc in out["source_documents"]:
            meta = doc.metadata
            loc_parts = [meta.get("source", "?")]
            if "sheet" in meta:
                loc_parts.append(f"sheet={meta['sheet']}")
            if "row" in meta:
                loc_parts.append(f"row={meta['row']}")
            print(f"  - {' | '.join(loc_parts)}: {doc.page_content[:80].replace(chr(10), ' ')}...")


if __name__ == "__main__":
    main()
